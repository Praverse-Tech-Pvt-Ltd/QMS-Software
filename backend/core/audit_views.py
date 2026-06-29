"""
Audit Trail Export — PDF and Excel.
GET /api/v1/audit-trail/?module=dms&record_id=<uuid>&date_from=&date_to=&format=pdf|excel
Only QA and Admin roles can access.
"""
import io
from django.http import HttpResponse
from django.utils.dateparse import parse_datetime
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.contrib.contenttypes.models import ContentType

from .models import AuditLog


def _get_logs(request):
    qs = AuditLog.objects.all().order_by('-timestamp')

    record_id = request.GET.get('record_id')
    module = request.GET.get('module')
    user_id = request.GET.get('user')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if record_id:
        qs = qs.filter(object_id=record_id)

    if module:
        # Filter by app label (module name maps to app)
        qs = qs.filter(content_type__app_label=module)

    if user_id:
        qs = qs.filter(user_id=user_id)

    if date_from:
        dt = parse_datetime(date_from)
        if dt:
            qs = qs.filter(timestamp__gte=dt)

    if date_to:
        dt = parse_datetime(date_to)
        if dt:
            qs = qs.filter(timestamp__lte=dt)

    return qs


class AuditTrailExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        role = getattr(request.user, 'role', '')
        if role not in ('Admin', 'QA', 'QA Head', 'QA Manager', 'QA Executive'):
            return Response({'error': 'Access denied. QA or Admin role required.'}, status=403)

        fmt = request.GET.get('format', 'json')
        logs = _get_logs(request)

        if fmt == 'pdf':
            return self._export_pdf(logs)
        elif fmt == 'excel':
            return self._export_excel(logs)
        else:
            data = [
                {
                    'id': log.id,
                    'user': log.user_email,
                    'action': log.action,
                    'record_type': str(log.content_type),
                    'record_id': log.object_id,
                    'timestamp': log.timestamp.isoformat(),
                    'reason': log.reason,
                    'changes': log.changes,
                }
                for log in logs[:500]
            ]
            return Response({'count': logs.count(), 'results': data})

    def _export_pdf(self, logs):
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
        except ImportError:
            return Response({'error': 'reportlab not installed'}, status=500)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=20)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph('QMS Audit Trail Export', styles['Title']))
        elements.append(Spacer(1, 12))

        data = [['Timestamp (UTC)', 'User', 'Action', 'Record Type', 'Record ID', 'Reason']]
        for log in logs[:1000]:
            data.append([
                log.timestamp.strftime('%Y-%m-%d %H:%M:%S UTC'),
                log.user_email or '',
                log.action,
                str(log.content_type),
                log.object_id,
                (log.reason or '')[:80],
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#dee2e6')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="audit_trail.pdf"'
        return response

    def _export_excel(self, logs):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return Response({'error': 'openpyxl not installed'}, status=500)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Audit Trail'

        headers = ['Timestamp (UTC)', 'User Email', 'Action', 'Record Type', 'Record ID', 'Reason', 'Changes']
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row_idx, log in enumerate(logs[:10000], 2):
            ws.cell(row=row_idx, column=1, value=log.timestamp.strftime('%Y-%m-%dT%H:%M:%SZ'))
            ws.cell(row=row_idx, column=2, value=log.user_email or '')
            ws.cell(row=row_idx, column=3, value=log.action)
            ws.cell(row=row_idx, column=4, value=str(log.content_type))
            ws.cell(row=row_idx, column=5, value=log.object_id)
            ws.cell(row=row_idx, column=6, value=log.reason or '')
            ws.cell(row=row_idx, column=7, value=str(log.changes or ''))

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="audit_trail.xlsx"'
        return response
